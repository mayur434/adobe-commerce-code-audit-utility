"""
Adobe Commerce Code Audit Report Generator
=============================================
Generates enterprise Excel report with charts from scanner findings.
"""

import re
import openpyxl
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict
from datetime import datetime

from lib.styles import (
    TITLE_FONT, SUBTITLE_FONT, HEADER_FONT, HEADER_FILL, HEADER_BORDER,
    SUMMARY_LABEL_FONT, SUMMARY_VALUE_FONT, BODY_FONT, BODY_FONT_BOLD,
    CODE_FONT, SECTION_FILL, THIN_BORDER, ZEBRA_FILL_1, ZEBRA_FILL_2,
    CENTER_ALIGN, CENTER_TOP, LEFT_TOP,
    severity_fill, severity_font, style_header_row,
    apply_zebra_and_borders, color_severity_col, color_priority_col,
)


class AuditReportGenerator:
    """Generates enterprise Excel report with charts from scanner findings."""

    def __init__(self, findings, stats, project_name, project_root):
        self.findings = findings
        self.stats = stats
        self.project_name = project_name
        self.project_root = project_root
        self.wb = openpyxl.Workbook()

    def generate(self, output_path):
        print("\n📊 Generating enterprise report...")
        self._sheet_executive_summary()

        # Detail sheets
        order = [
            # Code audit categories
            "Exception Handling", "Security", "Database", "Caching",
            "Code Structure", "Performance", "Deprecated", "Logging",
            "File Storage", "Reusability", "Test Coverage",
            "Dependency Injection", "Plugin Architecture", "Cron Jobs",
            "GraphQL", "Queue Processing", "Configuration", "Frontend Templates",
            "XML Configuration", "WebAPI & ACL", "DB Schema",
            "Infrastructure", "Cloud Deployment", "PHP Deep Analysis",
            "Event Observers", "Module Architecture", "Code Metrics",
            "Business Logic Identification",
            # DB dump analysis categories
            "DB: Table Structure", "DB: Index Analysis", "DB: Column Analysis",
            "DB: Foreign Keys", "DB: Naming Conventions", "DB: Storage Engine",
            "DB: Charset & Collation", "DB: Adobe Commerce Schema",
            "DB: Data Integrity", "DB: Performance",
        ]
        for cat in order:
            if cat in self.findings and self.findings[cat]:
                self._sheet_detail(cat, self.findings[cat])

        self._sheet_recommendations()
        self._sheet_action_plan()
        self._sheet_charts()

        self.wb.save(output_path)
        print(f"✅ Report: {output_path}")
        print(f"   Sheets: {len(self.wb.sheetnames)}")
        for s in self.wb.sheetnames:
            print(f"   - {s}")

    # ---------- Executive Summary ----------

    def _sheet_executive_summary(self):
        ws = self.wb.active
        ws.title = "Executive Summary"
        ws.sheet_properties.tabColor = "1F4E79"

        total = sum(self.stats.values())
        sev = {s: self.stats.get(s, 0) for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]}

        # --- Title Row (merged across A-F) ---
        title_text = f"{self.project_name} — ENTERPRISE CODE AUDIT REPORT"
        ws.merge_cells('A1:F1')
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = TITLE_FONT
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 36

        # --- Subtitle bar ---
        ws.merge_cells('A2:F2')
        ws.cell(row=2, column=1).fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        ws.row_dimensions[2].height = 4

        # --- Metadata rows ---
        meta = [
            ("Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Project Root:", self.project_root),
            ("Tool:", "Adobe Commerce Audit Utility v3.0 — Enterprise Deep Dive Edition"),
            ("Total Findings:", total),
        ]
        for i, (label, value) in enumerate(meta, 3):
            ws.cell(row=i, column=1, value=label).font = SUMMARY_LABEL_FONT
            ws.cell(row=i, column=1).alignment = LEFT_TOP
            val_cell = ws.cell(row=i, column=2, value=value)
            val_cell.font = SUMMARY_VALUE_FONT
            val_cell.alignment = LEFT_TOP
            ws.row_dimensions[i].height = 22
            if label == "Total Findings:":
                val_cell.font = Font(name='Calibri', bold=True, size=14, color='1F4E79')

        # --- Severity Breakdown Section ---
        sev_start = 8
        ws.merge_cells(f'A{sev_start}:F{sev_start}')
        sec = ws.cell(row=sev_start, column=1, value="SEVERITY BREAKDOWN")
        sec.font = SUBTITLE_FONT
        sec.fill = SECTION_FILL
        sec.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[sev_start].height = 28
        for c in range(1, 7):
            ws.cell(row=sev_start, column=c).fill = SECTION_FILL

        # Severity header
        sev_hdr_row = sev_start + 1
        for ci, h in enumerate(["Severity", "Count", "Description"], 1):
            cell = ws.cell(row=sev_hdr_row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = HEADER_BORDER

        # Severity data rows
        sev_data = [
            ("CRITICAL", sev["CRITICAL"], "Must fix immediately — security, data loss, production failures"),
            ("HIGH", sev["HIGH"], "Fix within 2 weeks — performance, reliability, architecture"),
            ("MEDIUM", sev["MEDIUM"], "Fix within 1 month — best practices, maintainability, deprecations"),
            ("LOW", sev["LOW"], "Backlog — code style, conventions, minor optimizations"),
            ("INFO", sev["INFO"], "Informational — no action required"),
        ]
        for i, (s, cnt, desc) in enumerate(sev_data, sev_hdr_row + 1):
            ws.cell(row=i, column=1, value=s).fill = severity_fill(s)
            ws.cell(row=i, column=1).font = severity_font(s)
            ws.cell(row=i, column=1).alignment = CENTER_TOP
            ws.cell(row=i, column=1).border = THIN_BORDER
            cnt_cell = ws.cell(row=i, column=2, value=cnt)
            cnt_cell.font = Font(name='Calibri', bold=True, size=12, color='333333')
            cnt_cell.alignment = CENTER_TOP
            cnt_cell.border = THIN_BORDER
            desc_cell = ws.cell(row=i, column=3, value=desc)
            desc_cell.font = BODY_FONT
            desc_cell.alignment = LEFT_TOP
            desc_cell.border = THIN_BORDER
            zebra = ZEBRA_FILL_1 if i % 2 == 0 else ZEBRA_FILL_2
            cnt_cell.fill = zebra
            desc_cell.fill = zebra
            ws.row_dimensions[i].height = 22

        # --- Category Breakdown Section ---
        cat_start = sev_hdr_row + len(sev_data) + 2
        ws.merge_cells(f'A{cat_start}:F{cat_start}')
        cs = ws.cell(row=cat_start, column=1, value="CATEGORY BREAKDOWN")
        cs.font = SUBTITLE_FONT
        cs.fill = SECTION_FILL
        cs.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[cat_start].height = 28
        for c in range(1, 7):
            ws.cell(row=cat_start, column=c).fill = SECTION_FILL

        cat_hdr_row = cat_start + 1
        cat_headers = ["Category", "Total", "Critical", "High", "Medium", "Low"]
        for ci, h in enumerate(cat_headers, 1):
            cell = ws.cell(row=cat_hdr_row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = HEADER_BORDER

        sorted_cats = sorted(self.findings.keys())
        for i, cat in enumerate(sorted_cats, cat_hdr_row + 1):
            items = self.findings[cat]
            zebra = ZEBRA_FILL_1 if (i - cat_hdr_row - 1) % 2 == 0 else ZEBRA_FILL_2
            ws.row_dimensions[i].height = 20
            row_data = [cat, len(items)]
            for s in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]:
                row_data.append(sum(1 for it in items if it["severity"] == s))
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=ci, value=val)
                cell.border = THIN_BORDER
                cell.fill = zebra
                if ci == 1:
                    cell.font = BODY_FONT_BOLD
                    cell.alignment = LEFT_TOP
                else:
                    cell.font = BODY_FONT
                    cell.alignment = CENTER_TOP
                    # Color non-zero severity counts
                    if ci >= 3 and val and val > 0:
                        sev_name = ["CRITICAL", "HIGH", "MEDIUM", "LOW"][ci - 3]
                        cell.fill = severity_fill(sev_name)
                        cell.font = severity_font(sev_name)

        # --- Top Risk Modules Section ---
        mod_counts = Counter()
        mod_crit = Counter()
        for items in self.findings.values():
            for item in items:
                mod_counts[item["module"]] += 1
                if item["severity"] == "CRITICAL":
                    mod_crit[item["module"]] += 1

        mod_start = cat_hdr_row + len(sorted_cats) + 2
        ws.merge_cells(f'A{mod_start}:F{mod_start}')
        ms = ws.cell(row=mod_start, column=1, value="TOP RISK MODULES")
        ms.font = SUBTITLE_FONT
        ms.fill = SECTION_FILL
        ms.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[mod_start].height = 28
        for c in range(1, 7):
            ws.cell(row=mod_start, column=c).fill = SECTION_FILL

        mod_hdr_row = mod_start + 1
        mod_headers = ["Module", "Total", "Critical", "Risk"]
        for ci, h in enumerate(mod_headers, 1):
            cell = ws.cell(row=mod_hdr_row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = HEADER_BORDER

        for i, (mod, cnt) in enumerate(mod_counts.most_common(15), mod_hdr_row + 1):
            c = mod_crit.get(mod, 0)
            risk = "🔴 Critical" if c > 3 else "🟠 High" if c > 0 else "🟡 Medium" if cnt > 5 else "🟢 Low"
            zebra = ZEBRA_FILL_1 if (i - mod_hdr_row - 1) % 2 == 0 else ZEBRA_FILL_2
            ws.row_dimensions[i].height = 20
            for ci, val in enumerate([mod, cnt, c, risk], 1):
                cell = ws.cell(row=i, column=ci, value=val)
                cell.border = THIN_BORDER
                cell.fill = zebra
                if ci == 1:
                    cell.font = BODY_FONT_BOLD
                    cell.alignment = LEFT_TOP
                elif ci == 4:
                    cell.font = BODY_FONT_BOLD
                    cell.alignment = CENTER_TOP
                else:
                    cell.font = BODY_FONT
                    cell.alignment = CENTER_TOP

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 70
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

    # ---------- Charts (last sheet) ----------

    def _sheet_charts(self):
        ws = self.wb.create_sheet("Charts")
        ws.sheet_properties.tabColor = "2E75B6"

        # ---- Data tables at top with proper formatting ----
        # Severity data: A1:B6
        sev_hdr_a = ws.cell(row=1, column=1, value="Severity")
        sev_hdr_a.font = HEADER_FONT; sev_hdr_a.fill = HEADER_FILL; sev_hdr_a.alignment = CENTER_ALIGN; sev_hdr_a.border = HEADER_BORDER
        sev_hdr_b = ws.cell(row=1, column=2, value="Count")
        sev_hdr_b.font = HEADER_FONT; sev_hdr_b.fill = HEADER_FILL; sev_hdr_b.alignment = CENTER_ALIGN; sev_hdr_b.border = HEADER_BORDER
        ws.row_dimensions[1].height = 24
        sev_order = ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]
        sev_colors = ["FF0000", "FF6600", "FFCC00", "92D050", "BDD7EE"]
        for i, s in enumerate(sev_order, 2):
            cell_s = ws.cell(row=i, column=1, value=s)
            cell_s.fill = severity_fill(s); cell_s.font = severity_font(s)
            cell_s.alignment = CENTER_TOP; cell_s.border = THIN_BORDER
            cell_v = ws.cell(row=i, column=2, value=self.stats.get(s, 0))
            cell_v.font = Font(name='Calibri', bold=True, size=11, color='333333')
            cell_v.alignment = CENTER_TOP; cell_v.border = THIN_BORDER
            ws.row_dimensions[i].height = 20

        # Module data: D1:E11
        mod_counts = Counter()
        for items in self.findings.values():
            for item in items:
                mod_counts[item["module"]] += 1
        top = mod_counts.most_common(10)
        for ci, hdr in [(4, "Module"), (5, "Findings")]:
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER_ALIGN; cell.border = HEADER_BORDER
        for i, (mod, cnt) in enumerate(top, 2):
            zebra = ZEBRA_FILL_1 if (i - 2) % 2 == 0 else ZEBRA_FILL_2
            cell_m = ws.cell(row=i, column=4, value=mod)
            cell_m.font = BODY_FONT; cell_m.alignment = LEFT_TOP; cell_m.border = THIN_BORDER; cell_m.fill = zebra
            cell_c = ws.cell(row=i, column=5, value=cnt)
            cell_c.font = Font(name='Calibri', bold=True, size=10, color='333333')
            cell_c.alignment = CENTER_TOP; cell_c.border = THIN_BORDER; cell_c.fill = zebra
            ws.row_dimensions[i].height = 20

        # Category data: G1:K(1+N)
        cats = sorted(self.findings.keys())
        for ci, hdr in enumerate(["Category", "Critical", "High", "Medium", "Low"], 7):
            cell = ws.cell(row=1, column=ci, value=hdr)
            cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER_ALIGN; cell.border = HEADER_BORDER
        for i, cat in enumerate(cats, 2):
            items = self.findings[cat]
            zebra = ZEBRA_FILL_1 if (i - 2) % 2 == 0 else ZEBRA_FILL_2
            cell_cat = ws.cell(row=i, column=7, value=cat)
            cell_cat.font = BODY_FONT; cell_cat.alignment = LEFT_TOP; cell_cat.border = THIN_BORDER; cell_cat.fill = zebra
            for offset, sev_name in enumerate(["CRITICAL", "HIGH", "MEDIUM", "LOW"], 8):
                val = sum(1 for x in items if x["severity"] == sev_name)
                cell = ws.cell(row=i, column=offset, value=val)
                cell.alignment = CENTER_TOP; cell.border = THIN_BORDER
                if val > 0:
                    cell.fill = severity_fill(sev_name); cell.font = severity_font(sev_name)
                else:
                    cell.font = BODY_FONT; cell.fill = zebra
            ws.row_dimensions[i].height = 20

        # Calculate where data ends to place charts below
        data_end_row = max(6, 1 + len(top), 1 + len(cats))
        chart_start = data_end_row + 3  # leave 2 blank rows

        # --- Chart 1: Severity Pie (left) ---
        pie = PieChart()
        pie.title = "Findings by Severity"
        pie.style = 10
        pie.width = 18; pie.height = 14
        labels = Reference(ws, min_col=1, min_row=2, max_row=6)
        vals = Reference(ws, min_col=2, min_row=1, max_row=6)
        pie.add_data(vals, titles_from_data=True)
        pie.set_categories(labels)
        for idx, color in enumerate(sev_colors):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = color
            pie.series[0].data_points.append(pt)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = True
        ws.add_chart(pie, f"A{chart_start}")

        # --- Chart 2: Top Modules Bar (right of pie) ---
        bar2 = BarChart()
        bar2.type = "bar"
        bar2.title = "Top 10 Modules by Finding Count"
        bar2.style = 10; bar2.width = 22; bar2.height = 14
        ml = Reference(ws, min_col=4, min_row=2, max_row=1 + len(top))
        md = Reference(ws, min_col=5, min_row=1, max_row=1 + len(top))
        bar2.add_data(md, titles_from_data=True)
        bar2.set_categories(ml)
        bar2.series[0].graphicalProperties.solidFill = "2E75B6"
        ws.add_chart(bar2, f"K{chart_start}")

        # --- Chart 3: Category Stacked Bar (below both) ---
        cat_chart_start = chart_start + 16  # pie/bar height ~14 rows + gap
        bar = BarChart()
        bar.type = "col"; bar.grouping = "stacked"
        bar.title = "Findings by Category & Severity"
        bar.y_axis.title = "Count"; bar.x_axis.title = "Category"
        bar.style = 10; bar.width = 36; bar.height = 18
        cat_labels = Reference(ws, min_col=7, min_row=2, max_row=1 + len(cats))
        for ci, clr in [(8, "FF0000"), (9, "FF6600"), (10, "FFCC00"), (11, "92D050")]:
            v = Reference(ws, min_col=ci, min_row=1, max_row=1 + len(cats))
            bar.add_data(v, titles_from_data=True)
            bar.series[-1].graphicalProperties.solidFill = clr
        bar.set_categories(cat_labels)
        ws.add_chart(bar, f"A{cat_chart_start}")

        # Column widths for data readability
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['G'].width = 28
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 10

    # ---------- Detail Sheets ----------

    def _sheet_detail(self, category, items):
        # Excel forbids  : \ / ? * [ ] in sheet names
        name = re.sub(r'[:\\/?\*\[\]]', '-', category)[:31]
        ws = self.wb.create_sheet(name)

        sevs = [i["severity"] for i in items]
        tab_color = "FF0000" if "CRITICAL" in sevs else "FF6600" if "HIGH" in sevs else "FFCC00" if "MEDIUM" in sevs else "92D050"
        ws.sheet_properties.tabColor = tab_color

        # --- Summary banner row ---
        crit_cnt = sum(1 for x in sevs if x == "CRITICAL")
        high_cnt = sum(1 for x in sevs if x == "HIGH")
        banner = f"{category}  —  {len(items)} findings  |  {crit_cnt} Critical  |  {high_cnt} High"
        ws.merge_cells('A1:J1')
        b_cell = ws.cell(row=1, column=1, value=banner)
        b_cell.font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
        b_cell.fill = PatternFill(start_color=tab_color, end_color=tab_color, fill_type='solid')
        b_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30
        for c in range(1, 11):
            ws.cell(row=1, column=c).fill = PatternFill(start_color=tab_color, end_color=tab_color, fill_type='solid')

        # --- Header row ---
        headers = ["#", "Module", "File Path", "Line #", "Issue Type", "Description", "Code Context", "Severity", "Recommendation", "Effort"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = HEADER_BORDER
        ws.row_dimensions[2].height = 28
        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f"A2:J2"

        # --- Data rows ---
        for idx, item in enumerate(items, 1):
            r = idx + 2
            ws.row_dimensions[r].height = 32
            ws.cell(row=r, column=1, value=idx)
            ws.cell(row=r, column=2, value=item["module"])
            ws.cell(row=r, column=3, value=item["file"])
            ws.cell(row=r, column=4, value=item["line"])
            ws.cell(row=r, column=5, value=item["type"])
            ws.cell(row=r, column=6, value=item["description"])
            code_cell = ws.cell(row=r, column=7, value=item["code"][:500])
            code_cell.font = CODE_FONT
            ws.cell(row=r, column=8, value=item["severity"])
            ws.cell(row=r, column=9, value=item["recommendation"])
            ws.cell(row=r, column=10, value=item["effort"])

        mr = len(items) + 2
        color_severity_col(ws, 8, mr)
        apply_zebra_and_borders(ws, mr, len(headers), data_start=3)

        # Style effort column as centered
        for r in range(3, mr + 1):
            ws.cell(row=r, column=10).alignment = CENTER_TOP

        widths = [6, 28, 55, 8, 28, 60, 55, 12, 65, 10]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---------- Recommendations ----------

    def _sheet_recommendations(self):
        ws = self.wb.create_sheet("Recommendations")
        ws.sheet_properties.tabColor = "00B050"

        headers = ["#", "Area", "Recommendation", "Expected Impact", "Effort", "Priority", "Details"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h)
        style_header_row(ws, len(headers))

        recs = [
            # Performance & CPU
            ["Performance", "Enable Redis for FPC + session + cache backend", "50-80% page load reduction, massive CPU savings", "Low", "P0",
             "bin/magento setup:config:set --cache-backend=redis --page-cache=redis --session-save=redis"],
            ["Performance", "Enable Varnish HTTP accelerator", "90%+ server load reduction for cacheable pages", "Medium", "P0",
             "Varnish 7.x as reverse proxy. Custom VCL for dynamic blocks. Set TTL per route."],
            ["Performance", "Fix all N+1 query patterns", "10-100x speedup on affected pages", "Medium", "P0",
             "Batch load collections before loops. addFieldToFilter('id', ['in' => $ids])."],
            ["Performance", "Add database indexes to custom tables", "50-90% faster queries on large tables", "Low", "P1",
             "Add btree indexes on sku, status, created_at, updated_at in db_schema.xml."],
            ["Performance", "Enable flat catalog for categories/products", "30-50% faster category pages", "Low", "P1",
             "Admin > Catalog > Storefront > Use Flat Catalog = Yes"],
            ["Performance", "Optimize cron scheduling", "Reduce CPU spikes and lock contention", "Low", "P1",
             "Stagger schedules. Add lock mechanisms. Use separate cron groups for heavy jobs."],
            ["Performance", "Enable OPcache preloading (PHP 8.2+)", "20-30% faster bootstrap", "Low", "P1",
             "Set opcache.preload in php.ini. Reduces compile overhead on every request."],

            # CPU Optimization
            ["CPU", "Split God classes (5+ files > 500 lines)", "Reduce memory per request, faster autoload", "High", "P1",
             "Break large files into focused services. Each class < 300 lines."],
            ["CPU", "Replace static LoggingHelper with DI logger", "Reduce static call overhead", "Medium", "P2",
             "Static calls prevent opcode optimization. Use Psr\\Log\\LoggerInterface."],
            ["CPU", "Minimize plugins on hot paths", "5-15% reduction in request time", "Medium", "P2",
             "Audit plugins on ProductRepository, QuoteRepository. Remove unnecessary around plugins."],
            ["CPU", "Optimize message queue consumers", "Better CPU distribution", "Medium", "P2",
             "Consolidate consumers. Set max-messages. Monitor memory per consumer."],

            # API Performance
            ["API", "Implement GraphQL response caching", "80% faster repeat queries", "Medium", "P0",
             "Add getIdentities() to resolvers. Use Varnish X-Magento-Tags."],
            ["API", "Add rate limiting to public endpoints", "Prevent abuse, protect resources", "Medium", "P1",
             "Configure webapi.xml throttle. Custom middleware for GraphQL rate limiting."],
            ["API", "Cache external API responses", "Eliminate redundant calls", "Medium", "P1",
             "TTL: payment config 5min, Cashify 1hr, Zoho tickets 5min."],
            ["API", "Enable response compression (gzip/brotli)", "40-70% smaller responses", "Low", "P1",
             "nginx: gzip_types application/json. Add brotli for modern browsers."],
            ["API", "Implement ETag/Last-Modified headers", "Eliminate transfers for unchanged data", "Medium", "P2",
             "Generate ETag from entity updated_at. Return 304 Not Modified."],

            # Logging (Disk Space)
            ["Logging", "Implement log rotation for all custom handlers", "Prevent disk exhaustion", "Low", "P0",
             "RotatingFileHandler maxFiles=7. Or /etc/logrotate.d/ for all custom log files."],
            ["Logging", "Gate debug/info logging behind config flag", "50-80% log volume reduction", "Low", "P1",
             "system.xml: Enable Debug Logging toggle. Check before ->debug() calls."],
            ["Logging", "Remove sensitive data from logs", "PCI-DSS / GDPR compliance", "Medium", "P0",
             "Mask card numbers, tokens, passwords. Use data masking utility."],
            ["Logging", "Consolidate to structured JSON logging", "Easier analysis, centralized monitoring", "High", "P2",
             "Monolog JSON formatter → ELK/CloudWatch. Keep only error logs on disk."],
            ["Logging", "Replace json_encode of full objects", "30-50% smaller log files", "Low", "P1",
             "Log only: ['order_id' => $id, 'status' => $status] instead of full objects."],

            # File Storage
            ["Storage", "Add cleanup cron for generated CSV/export files", "Prevent disk growth", "Low", "P1",
             "Purge files older than 7 days from var/export/, var/tmp/. Schedule daily."],
            ["Storage", "Configure S3 lifecycle policies", "Auto-cleanup old objects", "Low", "P1",
             "Transition to IA after 30 days, delete after 90 days for temp files."],
            ["Storage", "Stream large files instead of memory load", "Prevent OOM for large files", "Medium", "P1",
             "Use SplFileObject or fopen/fgets. Process in 1000-record batches."],

            # Security
            ["Security", "Implement CSP whitelist for all modules", "Prevent XSS from injected scripts", "Low", "P1",
             "Create etc/csp_whitelist.xml for each module with external domains."],
            ["Security", "Add webhook signature validation", "Prevent forged webhooks", "Medium", "P0",
             "Validate HMAC on all payment gateway webhooks. Reject unverified."],
            ["Security", "Remove all hardcoded credentials", "Prevent credential exposure", "Low", "P0",
             "Store secrets in Admin config with encrypted backend_model "
             "(Magento\\Config\\Model\\Config\\Backend\\Encrypted). Values are editable in "
             "Admin > Stores > Configuration without redeployment. For CI/CD pipelines, "
             "use bin/magento config:sensitive:set or config:set --lock-env."],

            # Code Quality
            ["Quality", "Add declare(strict_types=1) to all PHP files", "Catch type bugs at runtime", "Low", "P2",
             "Automated: find app/code -name '*.php' and add declaration."],
            ["Quality", "Enable PHPStan level 6+ in CI/CD", "Catch bugs before deploy", "Medium", "P1",
             "Install PHPStan + Magento extension. Start at level 4, increment."],
            ["Quality", "Add PHP_CodeSniffer with Magento2 standard", "Enforce code consistency", "Low", "P2",
             "magento/magento-coding-standard. Add pre-commit hook."],
            ["Quality", "Implement service contracts (interfaces)", "Enable testability/extensibility", "High", "P2",
             "Create Api/ interfaces for all services. Register in di.xml preference."],

            # Infrastructure
            ["Infra", "Separate Redis instances for cache/session/FPC", "Prevent conflicts, better failover", "Medium", "P1",
             "db0=cache, db1=FPC, db2=sessions. Set max-memory policies per instance."],
            ["Infra", "Enable slow query log", "Proactively find slow queries", "Low", "P1",
             "long_query_time=1, slow_query_log=ON. Review weekly."],
            ["Infra", "Health check endpoints for all integrations", "Early failure detection", "Medium", "P2",
             "Check: DB, Redis, OpenSearch, S3, external APIs."],
            ["Infra", "Configure auto-scaling for PaaS", "Handle traffic spikes", "Medium", "P1",
             "CPU threshold triggers. Define min/max instances. Load test."],

            # Deployment
            ["Deploy", "Add pre-deployment validation pipeline", "Catch issues before production", "Medium", "P1",
             "CI: PHPStan + PHPCS + Unit Tests + di:compile + static-content:deploy --dry-run"],
            ["Deploy", "Implement blue-green deployment", "Zero-downtime deployments", "High", "P2",
             "Maintain two identical environments. Switch traffic after validation."],

            # XML Configuration
            ["XML Config", "Add sortOrder to all plugins in di.xml", "Predictable plugin execution order", "Low", "P1",
             "Audit every <plugin> tag. Add sortOrder='10' to define explicit order."],
            ["XML Config", "Encrypt all sensitive system.xml fields", "Prevent credential exposure in DB", "Low", "P0",
             "Add backend_model='Magento\\Config\\Model\\Config\\Backend\\Encrypted' to password/secret fields."],
            ["XML Config", "Use config_path for cron schedules", "Admin-configurable cron timing", "Low", "P1",
             "Replace hardcoded <schedule> with <config_path> in crontab.xml."],
            ["XML Config", "Add <sequence> to all module.xml files", "Correct module load order", "Low", "P1",
             "List dependencies in <sequence> to prevent undefined class errors."],
            ["XML Config", "Move sandbox URLs from config.xml defaults", "Prevent production pointing to sandbox", "Low", "P0",
             "Set production endpoints as default. Override for dev/staging via env.php."],

            # WebAPI & ACL
            ["WebAPI", "Secure anonymous API endpoints", "Prevent unauthorized data access", "Low", "P0",
             "Change resource='anonymous' to resource='self' or custom ACL. Especially for POST/PUT/DELETE."],
            ["WebAPI", "Add rate limiting to all API endpoints", "DDoS/abuse protection", "Medium", "P0",
             "Implement nginx rate limiting + Magento webapi rate limiter for public endpoints."],
            ["WebAPI", "Add granular ACL for admin operations", "Principle of least privilege", "Medium", "P1",
             "Separate ACL: view, create, edit, delete, export for each module's admin area."],

            # DB Schema
            ["DB Schema", "Add indexes to custom tables", "50-90% faster filtered queries", "Low", "P0",
             "Index: status columns, foreign keys, created_at, updated_at, entity_id in db_schema.xml."],
            ["DB Schema", "Review wide tables for normalization", "Reduce row size, improve I/O", "High", "P2",
             "Tables with 20+ columns should be split into parent/detail or EAV pattern."],

            # Infrastructure
            ["Infra", "Add security headers to nginx", "Prevent clickjacking, XSS, MIME attacks", "Low", "P0",
             "X-Frame-Options, X-Content-Type-Options, Strict-Transport-Security, CSP, Referrer-Policy."],
            ["Infra", "Enable nginx gzip compression", "40-70% smaller responses", "Low", "P0",
             "gzip on; gzip_types text/css application/javascript application/json image/svg+xml;"],
            ["Infra", "Add nginx rate limiting", "Brute-force & DDoS protection", "Medium", "P0",
             "limit_req_zone for login, checkout, API. Example: rate=10r/s burst=20."],
            ["Infra", "Enable OPcache JIT (PHP 8.2+)", "20-40% CPU reduction for compute-heavy ops", "Low", "P1",
             "opcache.jit=1255, opcache.jit_buffer_size=256M in php.ini."],

            # Cloud Deployment
            ["Cloud", "Set SCD_STRATEGY=compact", "2-5x faster static content deployment", "Low", "P1",
             "Add to .magento.env.yaml stage.build. Compact generates only used themes/locales."],
            ["Cloud", "Reduce high-process consumers (50→10-20)", "Save 1-3GB RAM on cloud", "Low", "P0",
             "50 processes × 50MB = 2.5GB. Monitor throughput. Reduce to 10-20 with max_message=1000."],
            ["Cloud", "Monitor MySQL disk growth", "Prevent DB disk full outage", "Low", "P1",
             "5GB disk. Implement data retention policies. Archive old orders/logs/transactions."],

            # PHP Deep Analysis
            ["PHP", "Replace all MD5/SHA1 usage", "Eliminate weak cryptography", "Low", "P0",
             "Use hash('sha256', ...) or random_bytes(). MD5 collisions are computationally feasible."],
            ["PHP", "Fix all DateTime timezone issues", "Prevent date calculation bugs", "Low", "P0",
             "Always pass DateTimeZone to new DateTime(). Use Magento TimezoneInterface."],
            ["PHP", "Remove exit/die from non-CLI code", "Proper Magento request lifecycle", "Low", "P1",
             "Replace with throw Exception or return ResultInterface. exit() skips Magento shutdown."],
            ["PHP", "Implement HttpGet/PostActionInterface", "HTTP method enforcement on controllers", "Low", "P1",
             "All controllers should declare which HTTP methods they accept via interface."],
        ]

        for idx, rec in enumerate(recs, 1):
            r = idx + 1
            ws.cell(row=r, column=1, value=idx)
            for ci, val in enumerate(rec, 2):
                ws.cell(row=r, column=ci, value=val)

        mr = len(recs) + 1
        color_priority_col(ws, 6, mr)
        apply_zebra_and_borders(ws, mr, len(headers))

        # Color the Area column (col 2) with grouped colors
        area_colors = {
            "Performance": "4472C4", "CPU": "5B9BD5", "API": "2E75B6",
            "Logging": "ED7D31", "Storage": "FFC000",
            "Security": "FF0000", "Quality": "A5A5A5",
            "Infra": "70AD47", "Deploy": "44546A",
            "XML Config": "9B59B6", "WebAPI": "E74C3C",
            "DB Schema": "3498DB", "Cloud": "1ABC9C", "PHP": "F39C12",
        }
        for r in range(2, mr + 1):
            cell = ws.cell(row=r, column=2)
            area = str(cell.value) if cell.value else ''
            if area in area_colors:
                cell.fill = PatternFill(start_color=area_colors[area], end_color=area_colors[area], fill_type='solid')
                cell.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
                cell.alignment = CENTER_TOP

        widths = [6, 16, 55, 48, 10, 10, 85]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---------- Action Plan ----------

    def _sheet_action_plan(self):
        ws = self.wb.create_sheet("Action Plan")
        ws.sheet_properties.tabColor = "FFC000"

        # Banner
        ws.merge_cells('A1:G1')
        b = ws.cell(row=1, column=1, value="PRIORITIZED ACTION PLAN")
        b.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
        b.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        b.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30
        for c in range(1, 8):
            ws.cell(row=1, column=c).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

        headers = ["Priority", "Action Item", "Category", "Modules", "Effort", "Risk if Not Fixed", "Sprint"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = HEADER_BORDER
        ws.row_dimensions[2].height = 28
        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = "A2:G2"

        actions = []

        # P0 from criticals
        for cat, items in self.findings.items():
            for item in items:
                if item["severity"] == "CRITICAL":
                    actions.append([
                        "P0 - Immediate",
                        f"Fix: {item['type']} in {item['file']}:{item['line']}",
                        cat, item["module"], item["effort"],
                        item["description"][:60], "Sprint 1"
                    ])

        # P1 from grouped high
        high_by_type = defaultdict(list)
        for cat, items in self.findings.items():
            for item in items:
                if item["severity"] == "HIGH":
                    high_by_type[item["type"]].append(item)
        for itype, items in sorted(high_by_type.items(), key=lambda x: -len(x[1])):
            mods = ", ".join(sorted(set(i["module"] for i in items)))[:60]
            actions.append([
                "P1 - This Sprint",
                f"Fix {len(items)} {itype} issues", itype, mods,
                "Medium" if len(items) < 5 else "High",
                f"{len(items)} high-severity issues", "Sprint 2"
            ])

        # P2 from grouped medium
        med_by_type = defaultdict(list)
        for cat, items in self.findings.items():
            for item in items:
                if item["severity"] == "MEDIUM":
                    med_by_type[item["type"]].append(item)
        for itype, items in sorted(med_by_type.items(), key=lambda x: -len(x[1]))[:12]:
            mods = ", ".join(sorted(set(i["module"] for i in items)))[:60]
            actions.append([
                "P2 - Next Sprint",
                f"Address {len(items)} {itype} findings", itype, mods,
                "Medium", "Best practice violations", "Sprint 3-4"
            ])

        # P3/P4 static
        actions += [
            ["P3 - Backlog", "Add unit tests for critical modules", "Test Coverage", "All", "Very High", "Cannot validate correctness", "Sprint 4-6"],
            ["P3 - Backlog", "PHPStan level 6+ in CI/CD", "Quality", "All", "Medium", "Bugs reach production", "Sprint 4"],
            ["P3 - Backlog", "Set up log rotation", "Logging", "All", "Low", "Disk exhaustion", "Sprint 4"],
            ["P3 - Backlog", "Create shared Common module", "Reusability", "Multiple", "High", "Code duplication", "Sprint 5"],
            ["P4 - Nice to Have", "PHP 8.3 features adoption", "Modernization", "All", "Medium", "Technical debt", "Sprint 6+"],
        ]

        for idx, item in enumerate(actions, 1):
            r = idx + 2
            for ci, val in enumerate(item, 1):
                ws.cell(row=r, column=ci, value=val)

        mr = len(actions) + 2
        color_priority_col(ws, 1, mr)
        apply_zebra_and_borders(ws, mr, len(headers), data_start=3)

        # Style Sprint column centered
        for r in range(3, mr + 1):
            ws.cell(row=r, column=7).alignment = CENTER_TOP
            ws.cell(row=r, column=5).alignment = CENTER_TOP

        widths = [20, 58, 25, 42, 12, 52, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
